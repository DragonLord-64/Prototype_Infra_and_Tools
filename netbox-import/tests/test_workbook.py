import tempfile
import unittest
from pathlib import Path

import openpyxl

from netbox_import.workbook import WorkbookError, load_workbook


def _write_workbook(path, sheets):
    """sheets: {name: [header_row, *data_rows]}"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    wb.save(path)


class LoadWorkbookTests(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmpdir.cleanup)
        self.path = Path(self.tmpdir.name) / "inventory.xlsx"

    def test_reads_rows_as_dicts(self):
        _write_workbook(
            self.path,
            {"Sites": [["name", "slug", "status"], ["Lab", "lab", "active"]]},
        )
        sheets = load_workbook(self.path)
        self.assertEqual(sheets["Sites"], [{"name": "Lab", "slug": "lab", "status": "active"}])

    def test_ignores_unrecognized_sheets(self):
        _write_workbook(self.path, {"Notes": [["whatever"], ["x"]]})
        sheets = load_workbook(self.path)
        self.assertEqual(sheets, {})

    def test_skips_blank_rows(self):
        _write_workbook(
            self.path,
            {"Sites": [["name", "slug"], ["Lab", "lab"], [None, None], ["", ""]]},
        )
        sheets = load_workbook(self.path)
        self.assertEqual(len(sheets["Sites"]), 1)

    def test_missing_required_column_raises(self):
        _write_workbook(self.path, {"Racks": [["name"], ["Rack1"]]})  # missing "site"
        with self.assertRaises(WorkbookError):
            load_workbook(self.path)

    def test_strips_string_values(self):
        _write_workbook(
            self.path,
            {"Sites": [["name", "slug"], ["  Lab  ", "  lab  "]]},
        )
        sheets = load_workbook(self.path)
        self.assertEqual(sheets["Sites"][0]["name"], "Lab")
        self.assertEqual(sheets["Sites"][0]["slug"], "lab")

    def test_empty_sheet_is_empty_list(self):
        _write_workbook(self.path, {"Sites": [["name", "slug"]]})
        sheets = load_workbook(self.path)
        self.assertEqual(sheets["Sites"], [])


if __name__ == "__main__":
    unittest.main()
