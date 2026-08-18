"""Reads the inventory workbook into per-sheet lists of row dicts."""
from __future__ import annotations

import openpyxl

# Sheets are synced in this order -- later sheets reference objects created
# by earlier ones (e.g. Devices needs DeviceTypes/DeviceRoles/Sites/Racks
# to already exist). Any sheet not in this list is ignored; any sheet in
# this list that's missing from the workbook is simply skipped.
SHEET_ORDER = [
    "Manufacturers",
    "DeviceRoles",
    "DeviceTypes",
    "Sites",
    "Racks",
    "Devices",
    "VLANs",
    "Prefixes",
    "IPAddresses",
]

REQUIRED_COLUMNS = {
    "Manufacturers": ["name"],
    "DeviceRoles": ["name"],
    "DeviceTypes": ["manufacturer", "model"],
    "Sites": ["name"],
    "Racks": ["name", "site"],
    "Devices": ["name", "device_type", "role", "site"],
    "VLANs": ["vid", "name"],
    "Prefixes": ["prefix"],
    "IPAddresses": ["address"],
}


class WorkbookError(ValueError):
    pass


def load_workbook(path):
    """Returns {sheet_name: [row_dict, ...]} for every recognized sheet
    present in the workbook. Blank rows (no value in the sheet's first
    required column) are skipped."""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sheet_name in SHEET_ORDER:
        if sheet_name not in wb.sheetnames:
            continue
        result[sheet_name] = _read_sheet(wb[sheet_name], sheet_name)
    return result


def _read_sheet(ws, sheet_name):
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []

    missing = [c for c in REQUIRED_COLUMNS[sheet_name] if c not in header]
    if missing:
        raise WorkbookError(
            f"sheet {sheet_name!r} is missing required column(s): {', '.join(missing)}"
        )

    anchor = REQUIRED_COLUMNS[sheet_name][0]
    rows = []
    for raw in rows_iter:
        row = {k: v for k, v in zip(header, raw) if k}
        if row.get(anchor) in (None, ""):
            continue  # blank row
        rows.append({k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()})
    return rows
