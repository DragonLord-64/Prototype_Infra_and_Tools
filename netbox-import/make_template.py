#!/usr/bin/env python3
"""Regenerates template/inventory-template.xlsx. Run this after changing
the sheet/column schema in netbox_import/workbook.py or sync.py:

    python3 make_template.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

# (sheet name, [(column, example value, required)])
SCHEMA = [
    (
        "Manufacturers",
        [
            ("name", "Dell", True),
            ("slug", "", False),
        ],
    ),
    (
        "DeviceRoles",
        [
            ("name", "Server", True),
            ("slug", "", False),
            ("color", "2196f3", False),
        ],
    ),
    (
        "DeviceTypes",
        [
            ("manufacturer", "Dell", True),
            ("model", "PowerEdge R740", True),
            ("slug", "", False),
            ("u_height", 2, False),
        ],
    ),
    (
        "Sites",
        [
            ("name", "Lab", True),
            ("slug", "", False),
            ("status", "active", False),
        ],
    ),
    (
        "Racks",
        [
            ("name", "Rack1", True),
            ("site", "Lab", True),
            ("status", "active", False),
            ("u_height", 42, False),
        ],
    ),
    (
        "Devices",
        [
            ("name", "srv01", True),
            ("device_type", "PowerEdge R740", True),
            ("role", "Server", True),
            ("site", "Lab", True),
            ("rack", "Rack1", False),
            ("position", 10, False),
            ("face", "front", False),
            ("status", "active", False),
            ("serial", "", False),
        ],
    ),
    (
        "VLANs",
        [
            ("vid", 100, True),
            ("name", "servers", True),
            ("site", "Lab", False),
            ("status", "active", False),
        ],
    ),
    (
        "Prefixes",
        [
            ("prefix", "10.0.0.0/24", True),
            ("site", "Lab", False),
            ("status", "active", False),
            ("description", "", False),
        ],
    ),
    (
        "IPAddresses",
        [
            ("address", "10.0.0.10/24", True),
            ("device", "srv01", False),
            ("interface", "eth0", False),
            ("status", "active", False),
            ("primary", "true", False),
        ],
    ),
]

BOLD = Font(bold=True)


def build():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("ReadMe")
    readme["A1"] = "How to use this template"
    readme["A1"].font = Font(bold=True, size=14)
    readme.append([])
    readme.append(["Each tab below is one NetBox object type. Fill in rows, delete the"])
    readme.append(["example row, leave optional columns blank if you don't need them,"])
    readme.append(["then run: python -m netbox_import your-file.xlsx --dry-run"])
    readme.append([])
    readme.append(["Tab", "Required columns"])
    readme["A7"].font = BOLD
    readme["B7"].font = BOLD
    for sheet_name, columns in SCHEMA:
        required = ", ".join(c for c, _, req in columns if req)
        readme.append([sheet_name, required])
    readme.column_dimensions["A"].width = 40
    readme.column_dimensions["B"].width = 50

    for sheet_name, columns in SCHEMA:
        ws = wb.create_sheet(sheet_name)
        headers = [c for c, _, _ in columns]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = BOLD
        ws.append([example for _, example, _ in columns])
        for i, _ in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 20

    return wb


if __name__ == "__main__":
    out = Path(__file__).parent / "template" / "inventory-template.xlsx"
    out.parent.mkdir(exist_ok=True)
    build().save(out)
    print(f"wrote {out}")
